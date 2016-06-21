# !/usr/bin/python

import os
from openpyxl import Workbook, load_workbook


class Table:

    def __init__(self, fullpath, name):
        self.fullpath = fullpath
        self.name = name

#takes coordinates and woorksheet
#in the column specified, will get all the row values
# getRows(row, totalRow, columns, worksheet, ws)
def getRows(totalRow, columns, worksheet, outSheet, date):

    highRow = 1
    for i, col in enumerate(columns):
        row = 1
        c = worksheet.cell(row = row, column = col)
        s = outSheet.cell(row = totalRow, column = i+1)
        while c.value:
            print c.value
            print row
            print col
            s.value = c.value
            row += 1
            c = worksheet.cell(row = row, column = col)
            s = outSheet.cell(row = totalRow + row, column = i+1)
        if row > highRow:
            highRow = row

    return highRow


def main():

    docs = []
    for root, dirs, files in os.walk(".", topdown=False):
        for name in files:
            print(os.path.join(root, name))
            full_path = os.path.join(root, name)
            full_path = full_path[2:]
            if full_path.endswith(".xlsx"):
                #docs.append(full_path)
                d = Table(full_path, name)
                docs.append(d)
                #text_file.write("%s\n" % full_path)

    rawInput = raw_input('enter a list of terms: ')
    tokens = rawInput.split(' ')
    print tokens

    #new workbook to save results to
    outputBook = Workbook()
    ws = outputBook.active

    totalRow = 1
    for i, doc in enumerate(docs):
        date = doc.name.split(" ", 1)
        print "HERE IS THE DATE: " + str(date)
        fp = doc.fullpath
        workbook = load_workbook(fp)
        sheets = workbook.sheetnames
        worksheet = workbook[sheets[0]]


        #open the output sheet and put the labels there
        # c = 1
        # oc = ws.cell(row = 1, column = c)
        # for tok in tokens:
        #     oc.value = tok
        #     c += 1
        #     oc = ws.cell(row = 1, column = c)


        #set up controls for the loop
        row = 1
        col = 1
        d = worksheet.cell(row = row, column = col)

        columns = []
        #LOOP OVER COLUMN HEADER
        while d.value:
            #LOOP OVER TOKENS
            for i, tok in enumerate(tokens):
                if tok in str(d.value).lower() and "%" not in str(d.value).lower():
                    #add which column match was found
                    columns.append(col)
            #increment our stuff, move it along
            col += 1
            d = worksheet.cell(row = row, column = col)

        #now we have a small array containing the columns we need
        print "found match in columns: "
        print columns

        totalRow += getRows(totalRow, columns, worksheet, ws, date[0])


    outputBook.save('results.xlsx')


if __name__ == "__main__":
    main()
